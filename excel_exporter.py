import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict
import io


class ExcelExporter:
    """Export email data to Excel format"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_excel_from_emails(self, emails: List[Dict], filename: str = None) -> str:
        """Create Excel file from email data"""
        if not emails:
            raise ValueError("No emails to export")
        
        df = self._emails_to_dataframe(emails)
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"emails_export_{timestamp}.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        self._create_formatted_excel(df, filename)
        
        return filename
    
    def create_excel_buffer(self, emails: List[Dict]) -> io.BytesIO:
        """Create Excel file in memory buffer for download"""
        if not emails:
            raise ValueError("No emails to export")
        
        df = self._emails_to_dataframe(emails)
        
        buffer = io.BytesIO()
        self._create_formatted_excel(df, buffer)
        buffer.seek(0)
        
        return buffer
    
    def _emails_to_dataframe(self, emails: List[Dict]) -> pd.DataFrame:
        """Convert email list to pandas DataFrame"""
        data = []
        
        for email_data in emails:
            row = {
                'Date': email_data.get('date', ''),
                'From': email_data.get('from', ''),
                'To': email_data.get('to', ''),
                'Subject': email_data.get('subject', ''),
                'Body': email_data.get('body', ''),
                'Has Attachments': 'Yes' if email_data.get('has_attachments', False) else 'No',
                'Attachments': ', '.join(email_data.get('attachments', []))
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        return df
    
    def _create_formatted_excel(self, df: pd.DataFrame, output):
        """Create formatted Excel file with styling"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Emails"
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = self.worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        self._adjust_column_widths(df)
        
        self.worksheet.freeze_panes = "A2"
        
        if isinstance(output, str):
            self.workbook.save(output)
        else:
            self.workbook.save(output)
    
    def _adjust_column_widths(self, df: pd.DataFrame):
        """Adjust column widths based on content"""
        column_widths = {
            'Date': 20,
            'From': 30,
            'To': 30,
            'Subject': 40,
            'Body': 60,
            'Has Attachments': 15,
            'Attachments': 30
        }
        
        for idx, column in enumerate(df.columns, 1):
            column_letter = self.worksheet.cell(row=1, column=idx).column_letter
            width = column_widths.get(column, 15)
            self.worksheet.column_dimensions[column_letter].width = width
        
        for row in self.worksheet.iter_rows(min_row=2, max_row=self.worksheet.max_row):
            for cell in row:
                if cell.column_letter == self.worksheet.cell(row=1, column=5).column_letter:
                    if cell.value and len(str(cell.value)) > 500:
                        cell.value = str(cell.value)[:500] + "..."
    
    def create_summary_sheet(self, emails: List[Dict], filename: str):
        """Create Excel with summary and detailed sheets"""
        if not emails:
            raise ValueError("No emails to export")
        
        df = self._emails_to_dataframe(emails)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Email Details', index=False)
            
            summary_data = {
                'Metric': [
                    'Total Emails',
                    'Date Range',
                    'Unique Senders',
                    'Emails with Attachments',
                    'Export Date'
                ],
                'Value': [
                    len(emails),
                    f"{df['Date'].min()} to {df['Date'].max()}" if 'Date' in df.columns else 'N/A',
                    df['From'].nunique() if 'From' in df.columns else 0,
                    df[df['Has Attachments'] == 'Yes'].shape[0] if 'Has Attachments' in df.columns else 0,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            workbook = writer.book
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return filename
    
    def export_to_csv(self, emails: List[Dict], filename: str = None) -> str:
        """Export emails to CSV format"""
        if not emails:
            raise ValueError("No emails to export")
        
        df = self._emails_to_dataframe(emails)
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"emails_export_{timestamp}.csv"
        
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        df.to_csv(filename, index=False, encoding='utf-8-sig')
        
        return filename
